# reports/views.py
from collections import defaultdict

from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.utils import timezone, formats
from locations.models import MPKNumber
from pickups.models import PickupWasteBin
from users.models import Permission
from waste.models import WasteCost, WasteFraction
from .models import (
    MonthlyConfirmation,
    MonthlyConfirmationBin,
    SummaryCollectionSchedule,
)
from .forms import ReportFilterForm
from .services import import_collection_data, generate_mpk_cost_report
from django.views import View
from django.utils.decorators import method_decorator
from django.contrib.admin.views.decorators import staff_member_required
import datetime
from django.http import JsonResponse, HttpResponse
from django.db.models import Prefetch
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import json



def _odmien_rekord(n):
    if n == 1:
        return "1 rekord"
    elif 2 <= n <= 4:
        return f"{n} rekordy"
    else:
        return f"{n} rekordów"


@staff_member_required
def import_excel_view(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]

        if not excel_file.name.endswith((".xlsx", ".csv")):
            messages.error(
                request, "Nieprawidłowy format pliku. Proszę wgrać plik .xlsx lub .csv"
            )
            return redirect("reports:monthly_summary")

        try:
            results = import_collection_data(excel_file, request.user)

            parts = []
            if results["imported"] > 0:
                parts.append(f"Zaimportowano {_odmien_rekord(results['imported'])}.")
            if results["skipped"] > 0:
                parts.append(
                    f"{_odmien_rekord(results['skipped'])} pominięto – istnieją już w bazie."
                )
            if results["auto_confirmed"] > 0:
                parts.append(
                    f"Automatycznie potwierdzono {_odmien_rekord(results['auto_confirmed'])}."
                )

            if parts:
                messages.success(request, " ".join(parts))

            if results["errors"]:
                for error in results["errors"][
                    :5
                ]:  # Pokaż maksymalnie 5 pierwszych błędów
                    messages.error(request, error)
                if len(results["errors"]) > 5:
                    messages.error(
                        request, f"...oraz {len(results['errors']) - 5} innych błędów."
                    )

        except Exception as e:
            messages.error(request, f"Wystąpił błąd podczas importu: {str(e)}")

    return redirect("reports:monthly_summary")


@staff_member_required
def approve_confirmation(request, pk):
    confirmation = get_object_or_404(MonthlyConfirmation, pk=pk)

    if confirmation.status != "ZATWIERDZONE":
        confirmation.status = "ZATWIERDZONE"
        confirmation.approved_by = request.user
        confirmation.approved_at = timezone.now()
        confirmation.save()
        messages.success(
            request,
            f"Raport dla MPK {confirmation.mpk_number} został ostatecznie zatwierdzony.",
        )
    else:
        messages.warning(request, "Ten raport jest już zatwierdzony.")

    return redirect(
        reverse("reports:monthly_summary")
        + f"?month={confirmation.month.month}&year={confirmation.month.year}"
    )


@login_required
def monthly_summary_view(request):
    today = timezone.now().date()
    filter_status = request.GET.get('filter_status')

    form = ReportFilterForm(request.GET or None, user=request.user)
    
    selected_month = None
    selected_year = None
    selected_mpk = None
    
    if form.is_valid():
        selected_month = form.cleaned_data.get("month") or None
        selected_year = form.cleaned_data.get("year") or None
        selected_mpk = form.cleaned_data.get("mpk") or None

    # 1. Filtrowanie rekordów z harmonogramu
    query_filters = {}
    if selected_year: query_filters["year"] = int(selected_year)
    if selected_month: query_filters["month"] = int(selected_month)
    if selected_mpk: query_filters["mpk_number_id"] = selected_mpk
    
    if not request.user.is_superuser:
        allowed_ids = Permission.objects.filter(
            user=request.user, active=True
        ).values_list("mpk_number_id", flat=True)
        query_filters["mpk_number_id__in"] = allowed_ids

    records = SummaryCollectionSchedule.objects.filter(**query_filters).select_related(
        "mpk_number", "waste_fraction", "waste_fraction__fraction_type"
    ).prefetch_related("mpk_number__locations")

    # 2. Pobranie statusów potwierdzeń
    conf_filters = {}
    if selected_year: conf_filters["month__year"] = int(selected_year)
    if selected_month: conf_filters["month__month"] = int(selected_month)
    
    confirmations = MonthlyConfirmation.objects.filter(**conf_filters).values(
        "mpk_number_id", "month", "status"
    )
    status_map = {(c["mpk_number_id"], c["month"]): c["status"] for c in confirmations}

    # 3. Grupowanie danych i generowanie stylów dla kafelków
    grouped_data = {}
    for record in records:
        group_id = f"{record.mpk_number_id}_{record.year}_{record.month}"
        
        if group_id not in grouped_data:
            rec_date_obj = datetime.date(record.year, record.month, 1)
            current_status = status_map.get((record.mpk_number_id, rec_date_obj), "OCZEKUJE")
            
            # FILTRY (Pills)
            if filter_status == 'confirmed' and current_status not in ['ZATWIERDZONE', 'POTWIERDZONE']: continue
            if filter_status == 'action_required' and current_status in ['ZATWIERDZONE', 'POTWIERDZONE']: continue

            # Pobranie lokalizacji
            first_location = record.mpk_number.locations.first()
            loc_name = first_location.obj_name if first_location else "Brak nazwy"

            grouped_data[group_id] = {
                "mpk_id": record.mpk_number_id,
                "mpk_number": record.mpk_number.mpk_number,
                "location": loc_name,
                "year": record.year,
                "month": record.month,
                "month_name": formats.date_format(rec_date_obj, "F"),
                "status": current_status,
                "fractions": {},
            }
        
        wf = record.waste_fraction
        name = wf.fraction_type.name if wf.fraction_type else "Inne"
        capacity = getattr(wf, "capacity", "")
        capacity_str = f"{capacity}L" if capacity else ""
        
        # Unikalny klucz grupujący pojemniki tego samego typu w jednym boxie MPK
        group_key = f"{name}_{capacity}"

        # Jeśli danej frakcji jeszcze nie ma w kafelkach tego miesiąca - stwórz ją
        if group_key not in grouped_data[group_id]["fractions"]:
            lower_name = name.lower()
            if "zmieszane" in lower_name: icon, color = "bi-trash3-fill", "secondary"
            elif any(x in lower_name for x in ["plastik", "metal", "tworzywa"]): icon, color = "bi-recycle", "warning"
            elif any(x in lower_name for x in ["papier", "makulatura"]): icon, color = "bi-box-seam", "primary"
            elif "szk" in lower_name: icon, color = "bi-cup-straw", "success"
            elif "bio" in lower_name: icon, color = "bi-tree-fill", "success"
            else: icon, color = "bi-trash", "dark"

            grouped_data[group_id]["fractions"][group_key] = {
                "name": name,
                "capacity": capacity_str,
                "qty": 0,
                "icon": icon,
                "color": color,
            }
        
        # Zsumuj ilość zgłoszeń do wybranego kafelka
        grouped_data[group_id]["fractions"][group_key]["qty"] += record.quantity

    # 4. Formatowanie przed wysłaniem do HTML
    final_grouped_data = []
    for data in grouped_data.values():
        # Zamień słownik frakcji z powrotem na listę i posortuj
        data["fractions"] = sorted(data["fractions"].values(), key=lambda x: (x["name"], x["capacity"]))
        final_grouped_data.append(data)

    sorted_summary = sorted(final_grouped_data, key=lambda x: (-x["year"], -x["month"], x["mpk_number"]))

    return render(request, "reports/monthly_summary.html", {
        "form": form,
        "summary_data": sorted_summary,
        "selected_month": selected_month,
        "selected_year": selected_year,
    })


def _prepare_verification_data(mpk_id, month, year, confirmation):
    pickups = (
        PickupWasteBin.objects.filter(
            pickup__mpk_number_id=mpk_id,
            pickup__reported_at__month=month,
            pickup__reported_at__year=year,
        )
        .values("waste_fraction_id")
        .annotate(total=Sum("quantity"))
    )

    imports = (
        SummaryCollectionSchedule.objects.filter(
            mpk_number_id=mpk_id, month=month, year=year
        )
        .values("waste_fraction_id")
        .annotate(total=Sum("quantity"))
    )

    saved_bins = {b.waste_fraction_id: b for b in confirmation.bins.all()}
    all_fraction_ids = set(
        [p["waste_fraction_id"] for p in pickups]
        + [i["waste_fraction_id"] for i in imports]
    )
    fractions = WasteFraction.objects.filter(id__in=all_fraction_ids).select_related(
        "fraction_type"
    )

    comparison_data = []
    for f in fractions:
        reported = next(
            (p["total"] for p in pickups if p["waste_fraction_id"] == f.id), 0
        )
        collected = next(
            (i["total"] for i in imports if i["waste_fraction_id"] == f.id), 0
        )
        saved = saved_bins.get(f.id)

        comparison_data.append(
            {
                "fraction": f,
                "reported_qty": int(reported),
                "collected_qty": int(collected),
                "confirmed_qty": int(saved.confirmed_quantity)
                if saved
                else int(collected),
                "note": saved.note if saved else "",
                "is_conflict": reported != collected,
            }
        )

    return comparison_data


def _handle_verification_post(request, confirmation, comparison_data):
    decision = request.POST.get("decision_status")
    has_error = False
    temp_data = []

    # 1. Walidacja danych wejściowych
    diff_found = False
    for item in comparison_data:
        f = item["fraction"]
        collected_from_excel = item["collected_qty"]

        qty_input = request.POST.get(f"qty_{f.id}")
        note_input = request.POST.get(f"note_{f.id}", "").strip()

        if qty_input is not None:
            qty_confirmed = int(qty_input)

            # Sprawdzamy czy użytkownik zmienił wartość względem tego co podał dostawca w Excelu
            if qty_confirmed != int(collected_from_excel):
                diff_found = True
                if not note_input:
                    messages.error(
                        request,
                        f"Dla frakcji {f} wymagana jest uwaga przy rozbieżności!",
                    )
                    has_error = True

            temp_data.append({"fraction": f, "qty": qty_confirmed, "note": note_input})

    if decision == "KONFLIKT" and not diff_found:
        messages.error(
            request,
            "Wybrano status 'Występują rozbieżności', ale nie zmieniono żadnej ilości względem danych dostawcy!",
        )
        has_error = True

    if not has_error:
        for item in temp_data:
            MonthlyConfirmationBin.objects.update_or_create(
                confirmation=confirmation,
                waste_fraction=item["fraction"],
                defaults={"confirmed_quantity": item["qty"], "note": item["note"]},
            )
        if not confirmation.created_by:
            confirmation.created_by = request.user
        confirmation.updated_by = request.user
        confirmation.status = decision
        confirmation.approved_by = request.user
        confirmation.approved_at = timezone.now()
        confirmation.save()

        messages.success(request, "Weryfikacja została zapisana pomyślnie.")
        return redirect("reports:monthly_summary")

    return None


@login_required
def verification_view(request):
    today = timezone.now().date()
    month = int(request.GET.get("month", today.month))
    year = int(request.GET.get("year", today.year))
    mpk_id = request.GET.get("mpk")

    first_day = datetime.date(year, month, 1)

    if not mpk_id:
        return render(request, "reports/verification_form.html", {"no_mpk": True})

    confirmation, created = MonthlyConfirmation.objects.get_or_create(
        mpk_number_id=mpk_id, month=first_day
    )

    pickups = PickupWasteBin.objects.filter(
        pickup__mpk_number_id=mpk_id,
        pickup__reported_at__month=month,
        pickup__reported_at__year=year
    ).values('waste_fraction_id').annotate(total=Sum('quantity'))
    
    imports = SummaryCollectionSchedule.objects.filter(
        mpk_number_id=mpk_id,
        month=month,
        year=year
    ).values('waste_fraction_id').annotate(total=Sum('quantity'))

    saved_bins = {b.waste_fraction_id: b for b in confirmation.bins.all()}
    all_fraction_ids = set([p['waste_fraction_id'] for p in pickups] + [i['waste_fraction_id'] for i in imports])
    fractions = WasteFraction.objects.filter(id__in=all_fraction_ids).select_related('fraction_type')

    comparison_data = []
    for f in fractions:
        reported = next((p['total'] for p in pickups if p['waste_fraction_id'] == f.id), 0)
        collected = next((i['total'] for i in imports if i['waste_fraction_id'] == f.id), 0)
        saved = saved_bins.get(f.id)
        
        comparison_data.append({
            'fraction': f,
            'reported_qty': int(reported),
            'collected_qty': int(collected),
            'confirmed_qty': int(saved.confirmed_quantity) if saved else int(collected),
            'note': saved.note if saved else "",
            'is_conflict': reported != collected
        })

    if request.method == 'POST':
        decision = request.POST.get('decision_status')
        has_error = False
        temp_data = []

        # 1. Walidacja danych wejściowych
        diff_found = False
        for f in fractions:

            _reported_in_system = next((p['total'] for p in pickups if p['waste_fraction_id'] == f.id), 0)
            collected_from_excel = next((i['total'] for i in imports if i['waste_fraction_id'] == f.id), 0)
            
            qty_input = request.POST.get(f'qty_{f.id}')
            note_input = request.POST.get(f'note_{f.id}', "").strip()

            if qty_input is not None:
                qty_confirmed = int(qty_input)
                
                # Sprawdzamy czy użytkownik zmienił wartość względem tego co podał dostawca w Excelu
                if qty_confirmed != int(collected_from_excel):
                    diff_found = True
                    if not note_input:
                        messages.error(request, f"Dla frakcji {f} wymagana jest uwaga przy rozbieżności!")
                        has_error = True
                
                temp_data.append({
                    'fraction': f,
                    'qty': qty_confirmed,
                    'note': note_input
                })

        if decision == 'KONFLIKT' and not diff_found:
            messages.error(request, "Wybrano status 'Występują rozbieżności', ale nie zmieniono żadnej ilości względem danych dostawcy!")
            has_error = True

        if not has_error:
            existing_bins = {
                b.waste_fraction_id: b
                for b in MonthlyConfirmationBin.objects.filter(confirmation=confirmation)
            }
            to_create = []
            to_update = []

            for item in temp_data:
                fraction_id = item['fraction'].id
                if fraction_id in existing_bins:
                    bin_obj = existing_bins[fraction_id]
                    if bin_obj.confirmed_quantity != item['qty'] or bin_obj.note != item['note']:
                        bin_obj.confirmed_quantity = item['qty']
                        bin_obj.note = item['note']
                        to_update.append(bin_obj)
                else:
                    to_create.append(
                        MonthlyConfirmationBin(
                            confirmation=confirmation,
                            waste_fraction=item['fraction'],
                            confirmed_quantity=item['qty'],
                            note=item['note']
                        )
                    )

            if to_create:
                MonthlyConfirmationBin.objects.bulk_create(to_create)
            if to_update:
                MonthlyConfirmationBin.objects.bulk_update(to_update, ['confirmed_quantity', 'note'])

            if not confirmation.created_by:
                confirmation.created_by = request.user
            confirmation.updated_by = request.user
            confirmation.status = decision
            confirmation.approved_by = request.user
            confirmation.approved_at = timezone.now()
            confirmation.save()
            
            messages.success(request, "Weryfikacja została zapisana pomyślnie.")
            return redirect('reports:monthly_summary')

    return render(request, 'reports/verification_form.html', {
        'confirmation': confirmation,
        'comparison_data': comparison_data,
        'month': month,
        'year': year
    })


@staff_member_required
def edit_summaries_view(request):
    year_param = request.GET.get('year')
    month_param = request.GET.get('month')
    mpk = request.GET.get('mpk')
    status_filter = request.GET.get('status')
    
    if not request.GET:
        year = str(timezone.now().year)
        month = str(timezone.now().month)
    else:
        year = year_param
        month = month_param

    queryset = SummaryCollectionSchedule.objects.select_related(
        'mpk_number', 'waste_fraction', 'waste_fraction__fraction_type'
    )

    if year:
        try:
            queryset = queryset.filter(year=int(year))
        except ValueError:
            pass

    if month:
        try:
            queryset = queryset.filter(month=int(month))
        except ValueError:
            pass
    if mpk:
        queryset = queryset.filter(mpk_number__mpk_number__icontains=mpk)
        
    conf_qs = MonthlyConfirmation.objects.prefetch_related('bins')
    if year:
        conf_qs = conf_qs.filter(month__year=int(year))
    if month:
        conf_qs = conf_qs.filter(month__month=int(month))

    confirmations_prefetch = Prefetch(
        'mpk_number__confirmations',
        queryset=conf_qs,
        to_attr='all_confirmations' 
    )
    
    queryset = queryset.prefetch_related(confirmations_prefetch)

    records = list(queryset)

    for record in records:
        record.confirmation_status = None
        record.confirmation_note = None
        
        if hasattr(record.mpk_number, 'all_confirmations') and record.mpk_number.all_confirmations:
            conf = next((c for c in record.mpk_number.all_confirmations if c.month.year == record.year and c.month.month == record.month), None)
            
            if conf:
                record.confirmation_status = conf.status
                
                for bin in conf.bins.all():
                    if bin.waste_fraction_id == record.waste_fraction_id:
                        record.confirmation_note = bin.note
                        break

    if status_filter:
        records = [r for r in records if r.confirmation_status == status_filter or (status_filter == 'BRAK' and r.confirmation_status is None)]

    available_years = SummaryCollectionSchedule.objects.values_list('year', flat=True).distinct().order_by('-year')
    available_months = range(1, 13)
    available_statuses = MonthlyConfirmation.STATUS_CHOICES

    context = {
        'records': records,
        'selected_year': int(year) if year else None,
        'selected_month': int(month) if month else None,
        'selected_mpk': mpk,
        'selected_status': status_filter,
        'available_years': available_years,
        'available_months': available_months,
        'available_statuses': available_statuses,
    }
    return render(request, 'reports/edit_summaries.html', context)

def update_summary_quantity(request):
    if not (request.user.is_authenticated and request.user.is_staff and request.user.is_active):
        return JsonResponse({'status': 'error', 'message': 'Brak uprawnień'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            record_id = data.get('id')
            new_quantity = data.get('quantity')

            record = get_object_or_404(SummaryCollectionSchedule, id=record_id)
            record.quantity = int(new_quantity)
            record.save()
            return JsonResponse({'status': 'success', 'new_quantity': record.quantity})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@staff_member_required
def export_summaries_xlsx(request):
    year = request.GET.get('year')
    month = request.GET.get('month')
    mpk = request.GET.get('mpk')
    status_filter = request.GET.get('status')

    queryset = SummaryCollectionSchedule.objects.select_related(
        'mpk_number', 'waste_fraction', 'waste_fraction__fraction_type'
    )

    if year is None and month is None:
        year = str(timezone.now().year)
        month = str(timezone.now().month)
        
    if year and year.isdigit():
        queryset = queryset.filter(year=int(year))

    if month and month.isdigit():
        queryset = queryset.filter(month=int(month))
    
    if mpk:
        queryset = queryset.filter(mpk_number__mpk_number__icontains=mpk)

    y = int(year) if year and str(year).isdigit() else timezone.now().year
    m = int(month) if month and str(month).isdigit() else timezone.now().month
    
    confirmations_prefetch = Prefetch(
        'mpk_number__confirmations',
        queryset=MonthlyConfirmation.objects.filter(
            month__year=y,
            month__month=m
        ).prefetch_related('bins'),
        to_attr='current_confirmation'
    )
    queryset = queryset.prefetch_related(confirmations_prefetch)

    records = list(queryset)

    for record in records:
        record.confirmation_status = None
        record.confirmation_note = None
        if hasattr(record.mpk_number, 'current_confirmation') and record.mpk_number.current_confirmation:
            conf = record.mpk_number.current_confirmation[0]
            record.confirmation_status = conf.status
            for bin in conf.bins.all():
                if bin.waste_fraction_id == record.waste_fraction_id:
                    record.confirmation_note = bin.note
                    break

    if status_filter:
        records = [r for r in records if r.confirmation_status == status_filter or (status_filter == 'BRAK' and r.confirmation_status is None)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zestawienia"

    fraction_cols = set()
    for record in records:
        fraction_str = f"{record.waste_fraction.fraction_type.name} - {record.waste_fraction.capacity}{record.waste_fraction.unit or 'L'}"
        fraction_cols.add(fraction_str)

    fraction_cols = sorted(list(fraction_cols))

    headers = ['Numer MPK', 'Rok', 'Miesiąc'] + fraction_cols + ['Status Akceptacji', 'Uwagi MPK']
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    attention_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    approved_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    from collections import defaultdict
    pivot_data = defaultdict(dict)
    mpk_metadata = defaultdict(lambda: {'status': None, 'notes': []})

    for record in records:
        key = (record.mpk_number.mpk_number, record.year, record.month)
        fraction_str = f"{record.waste_fraction.fraction_type.name} - {record.waste_fraction.capacity}{record.waste_fraction.unit or 'L'}"

        pivot_data[key][fraction_str] = {
            'quantity': record.quantity,
            'status': record.confirmation_status,
            'note': record.confirmation_note
        }

        # Track status and notes per MPK row
        if record.confirmation_status:
            mpk_metadata[key]['status'] = record.confirmation_status
        if record.confirmation_note:
            mpk_metadata[key]['notes'].append(f"[{fraction_str}] {record.confirmation_note}")

    for row_num, (key, row_dict) in enumerate(pivot_data.items(), 2):
        row_data = [key[0], key[1], key[2]]

        for frac_col in fraction_cols:
            if frac_col in row_dict:
                row_data.append(row_dict[frac_col]['quantity'])
            else:
                row_data.append(0)

        meta = mpk_metadata[key]
        row_data.append(meta['status'] or 'Brak')
        row_data.append(' | '.join(meta['notes']) if meta['notes'] else '')

        ws.append(row_data)

        for frac_idx, frac_col in enumerate(fraction_cols):
            col_num = 4 + frac_idx # Offset by 3 base columns
            if frac_col in row_dict:
                cell_data = row_dict[frac_col]
                cell = ws.cell(row=row_num, column=col_num)

                if cell_data['status'] == 'ZATWIERDZONE':
                    cell.fill = approved_fill
                elif cell_data['status'] == 'KONFLIKT' or cell_data['note']:
                    cell.fill = attention_fill

        status_col_num = 4 + len(fraction_cols)
        notes_col_num = 5 + len(fraction_cols)


        if meta["status"] == "ZATWIERDZONE" or "Zgodność automatyczna" in meta["notes"]:
            ws.cell(row=row_num, column=status_col_num).fill = approved_fill
            ws.cell(row=row_num, column=notes_col_num).fill = approved_fill
        elif meta['status'] == 'KONFLIKT' or meta['notes']:
            ws.cell(row=row_num, column=status_col_num).fill = attention_fill
            ws.cell(row=row_num, column=notes_col_num).fill = attention_fill


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="zestawienia_{y}_{m}.xlsx"'
    wb.save(response)
    return response


@method_decorator(staff_member_required, name='dispatch')
class ExportCostReportView(View):
    def get(self, request, *args, **kwargs):
        year_str = request.GET.get('year', '')
        month_str = request.GET.get('month', '')
        mpk_str = request.GET.get('mpk_number_id', '')
        fmt = request.GET.get('format', 'xlsx')

        year = int(year_str) if year_str.isdigit() else None
        month = int(month_str) if month_str.isdigit() else None
        mpk_number_id = int(mpk_str) if mpk_str.isdigit() else None

        if fmt not in ['xlsx', 'csv']:
            fmt = 'xlsx'

        file_content = generate_mpk_cost_report(year=year, month=month, mpk_number_id=mpk_number_id, report_format=fmt)

        if fmt == 'csv':
            response = HttpResponse(file_content, content_type='text/csv')
            filename = 'raport_kosztowy.csv'
        else:
            response = HttpResponse(file_content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = 'raport_kosztowy.xlsx'

        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

def _fraction_css(name: str) -> str:
    lower = name.lower()
    if 'zmieszane' in lower or 'zmieszany' in lower:
        return 'fraction-badge--mixed'
    if 'papier' in lower or 'makulatura' in lower:
        return 'fraction-badge--paper'
    if 'plastik' in lower or 'tworzywa' in lower or 'metal' in lower:
        return 'fraction-badge--plastic'
    if 'szkł' in lower or 'szk' in lower:
        return 'fraction-badge--glass'
    if 'bio' in lower:
        return 'fraction-badge--bio'
    return 'fraction-badge--default'

MONTH_NAMES = {
    1: 'Styczeń', 2: 'Luty', 3: 'Marzec', 4: 'Kwiecień',
    5: 'Maj', 6: 'Czerwiec', 7: 'Lipiec', 8: 'Sierpień',
    9: 'Wrzesień', 10: 'Październik', 11: 'Listopad', 12: 'Grudzień',
}

MONTH_CHOICES = [(i, MONTH_NAMES[i]) for i in range(1, 13)]
 
 
@login_required
def cost_summary_view(request):
    """
    Zestawienie kosztowe: SummaryCollectionSchedule × WasteCost.
    """
    today = timezone.now().date()
 
    # ── Odczyt filtrów z GET ─────────────────────────────────────────
    year_str  = request.GET.get('year', '').strip()
    month_str = request.GET.get('month', '').strip()
    mpk_str   = request.GET.get('mpk_number_id', '').strip()
 
    selected_year     = int(year_str)  if year_str.isdigit()  else None
    selected_month    = int(month_str) if month_str.isdigit() else None
    selected_mpk_id   = int(mpk_str)   if mpk_str.isdigit()   else None
 
    # Domyślnie: bieżący rok
    if selected_year is None:
        selected_year = today.year
 
    # ── Filtracja zestawień odbioru ──────────────────────────────────
    qs = SummaryCollectionSchedule.objects.select_related(
        'mpk_number', 'waste_fraction', 'waste_fraction__fraction_type'
    )
 
    if selected_year:
        qs = qs.filter(year=selected_year)
    if selected_month:
        qs = qs.filter(month=selected_month)
    if selected_mpk_id:
        qs = qs.filter(mpk_number_id=selected_mpk_id)
 
    # Ogranicz do MPK, do których użytkownik ma dostęp
    if not request.user.is_superuser:
        allowed_ids = Permission.objects.filter(
            user=request.user, active=True
        ).values_list('mpk_number_id', flat=True)
        qs = qs.filter(mpk_number_id__in=allowed_ids)
 
    summaries = list(qs)
 
    # ── Pobranie wszystkich stawek do pamięci ────────────────────────
    all_costs = list(WasteCost.objects.all().order_by('-date_from'))
 
    def get_unit_cost(fraction_id, target_date):
        for cost in all_costs:
            if cost.waste_fraction_id == fraction_id:
                if cost.date_from <= target_date and (
                    cost.date_to is None or cost.date_to >= target_date
                ):
                    return cost.cost
        return None
 
    # ── Budowanie wierszy tabeli ─────────────────────────────────────
    rows = []
    grand_total = 0
    missing_costs = []
 
    # Struktury do wykresu i sidebara
    chart_values   = defaultdict(lambda: defaultdict(float))  # mpk -> fraction -> koszt
    mpk_totals     = defaultdict(float)                        # mpk -> łączny koszt
    fraction_totals_map = defaultdict(float)                   # fraction_name -> łączny koszt
 
    for s in summaries:
        target_date = s.date_summary
        unit_cost   = get_unit_cost(s.waste_fraction_id, target_date)
        total_cost  = (unit_cost * s.quantity) if unit_cost is not None else None
 
        fraction_name = s.waste_fraction.fraction_type.name
        capacity      = s.waste_fraction.capacity
        unit          = s.waste_fraction.unit or 'szt'
        fraction_label = f"{fraction_name} ({capacity} {unit})"
        
        row = {
            'year':          s.year,
            'month':         s.month,
            'month_name':    MONTH_NAMES.get(s.month, str(s.month)),
            'mpk_number':    s.mpk_number.mpk_number,
            'fraction_name': fraction_name,
            'fraction_label': fraction_label,
            'fraction_css':  _fraction_css(fraction_name),
            'capacity':       capacity,
            'unit':           unit,
            'quantity':      s.quantity,
            'unit_cost':     unit_cost,
            'total_cost':    total_cost,
        }
        rows.append(row)
 
        if unit_cost is None:
            missing_costs.append(row)
        else:
            grand_total += float(total_cost)
            chart_values[s.mpk_number.mpk_number][fraction_name] += float(total_cost)
            mpk_totals[s.mpk_number.mpk_number] += float(total_cost)
            fraction_totals_map[fraction_name] += float(total_cost)
 
    # ── Dane dla Chart.js ────────────────────────────────────────────
    all_mpks      = sorted(chart_values.keys())
    all_fractions = sorted({fn for mpk_data in chart_values.values() for fn in mpk_data})
 
    chart_data_json = json.dumps({
        'mpks':      all_mpks,
        'fractions': all_fractions,
        'values':    {
            str(mpk): {frac: chart_values[mpk].get(frac, 0) for frac in all_fractions}
            for mpk in all_mpks
        },
    })
 
    # ── Top MPK ──────────────────────────────────────────────────────
    top_mpks = sorted(
        [{'mpk_number': k, 'total': v} for k, v in mpk_totals.items()],
        key=lambda x: x['total'],
        reverse=True,
    )[:5]
 
    # ── Koszty wg frakcji (sidebar) ──────────────────────────────────
    fraction_totals = sorted(
        [
            {'name': k, 'total': v, 'css': _fraction_css(k)}
            for k, v in fraction_totals_map.items()
        ],
        key=lambda x: x['total'],
        reverse=True,
    )
 
    # ── Dane pomocnicze do filtrów ───────────────────────────────────
    if request.user.is_superuser:
        mpk_numbers = MPKNumber.objects.filter(active=True).order_by('mpk_number')
    else:
        allowed_ids = Permission.objects.filter(
            user=request.user, active=True
        ).values_list('mpk_number_id', flat=True)
        mpk_numbers = MPKNumber.objects.filter(id__in=allowed_ids, active=True).order_by('mpk_number')
 
    available_years = (
        SummaryCollectionSchedule.objects
        .values_list('year', flat=True)
        .distinct()
        .order_by('-year')
    )
 
    context = {
        # Dane tabeli
        'rows':           rows,
        'missing_costs':  missing_costs,
        'grand_total':    grand_total,
 
        # Wykres
        'chart_data_json': chart_data_json,
 
        # Sidebar
        'top_mpks':         top_mpks,
        'fraction_totals':  fraction_totals,
        'active_mpk_count': len(mpk_totals),
        'fraction_count':   len(fraction_totals_map),
 
        # Filtry
        'available_years':   available_years,
        'months':            MONTH_CHOICES,
        'mpk_numbers':       mpk_numbers,
        'selected_year':     selected_year,
        'selected_month':    selected_month,
        'selected_mpk_id':   selected_mpk_id,
        'selected_month_name': MONTH_NAMES.get(selected_month, '') if selected_month else '',
    }
 
    return render(request, 'reports/cost_summary.html', context)