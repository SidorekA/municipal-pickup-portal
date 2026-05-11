import io
from django.test import TestCase
from django.contrib.auth import get_user_model
from django.utils import timezone
from openpyxl import load_workbook

from pickups.models import Pickup, PickupWasteBin
from locations.models import Location, MPKNumber
from waste.models import WasteFraction, WasteFractionType
from pickups.excel_generator import generate_pickup_excel, HEADERS

User = get_user_model()

class ExcelGeneratorTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="Reporter",
            first_name="Jan",
            last_name="Kowalski",
            password="testpassword"
        )
        self.mpk = MPKNumber.objects.create(mpk_number=1234)
        self.location = Location.objects.create(
            obj_name="Testowy Obiekt",
            org_unit_name="Jednostka Testowa",
            localization="Adres Testowy 12",
            mpk_number=self.mpk,
        )

        self.pickup = Pickup.objects.create(
            location=self.location,
            mpk_number=self.mpk,
            reporter=self.user,
            contact_phone="123456789",
            note="Testowa notatka",
            reported_at=timezone.datetime(2023, 10, 25, 12, 30, tzinfo=timezone.get_current_timezone())
        )

        # Create a matching fraction
        self.fraction_type_zmieszane = WasteFractionType.objects.create(
            code=200301,
            name="Zmieszane"
        )
        self.fraction_zmieszane_1100 = WasteFraction.objects.create(
            fraction_type=self.fraction_type_zmieszane,
            capacity=1100,
            unit="l"
        )

        PickupWasteBin.objects.create(
            pickup=self.pickup,
            waste_fraction=self.fraction_zmieszane_1100,
            quantity=3
        )

        # Create a second matching fraction for a capacityless match (bio)
        self.fraction_type_bio = WasteFractionType.objects.create(
            code=200201,
            name="Bio"
        )
        self.fraction_bio = WasteFraction.objects.create(
            fraction_type=self.fraction_type_bio,
            capacity=240, # should match just by name for bio
            unit="l"
        )

        PickupWasteBin.objects.create(
            pickup=self.pickup,
            waste_fraction=self.fraction_bio,
            quantity=1
        )

    def test_generate_pickup_excel_returns_bytes(self):
        result = generate_pickup_excel(self.pickup)
        self.assertIsInstance(result, bytes)

    def test_generate_pickup_excel_content(self):
        result = generate_pickup_excel(self.pickup)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        self.assertEqual(ws.title, "zgloszenia_odpadow_do_wysylania")

        # Check headers
        for col_idx, header in enumerate(HEADERS, 1):
            self.assertEqual(ws.cell(row=1, column=col_idx).value, header)

        # Check row 2 (data)
        # Find index for some fields
        mpk_idx = HEADERS.index('numer_mpk') + 1
        org_idx = HEADERS.index('nazwa_komorki_organizacyjnej') + 1
        obj_idx = HEADERS.index('nazwa_obiektu') + 1
        loc_idx = HEADERS.index('lokalizacja') + 1
        phone_idx = HEADERS.index('numer_telefonu') + 1
        date_idx = HEADERS.index('Utworzony') + 1
        user_idx = HEADERS.index('Utworzone przez') + 1
        note_idx = HEADERS.index('informacje_dodatkowe') + 1

        self.assertEqual(ws.cell(row=2, column=mpk_idx).value, 1234)
        self.assertEqual(ws.cell(row=2, column=org_idx).value, "Jednostka Testowa")
        self.assertEqual(ws.cell(row=2, column=obj_idx).value, "Testowy Obiekt")
        self.assertEqual(ws.cell(row=2, column=loc_idx).value, "Adres Testowy 12")
        self.assertEqual(ws.cell(row=2, column=phone_idx).value, "123456789")
        self.assertEqual(ws.cell(row=2, column=date_idx).value, self.pickup.reported_at.strftime("%Y-%m-%d %H:%M"))
        self.assertEqual(ws.cell(row=2, column=user_idx).value, "Jan Kowalski")
        self.assertEqual(ws.cell(row=2, column=note_idx).value, "Testowa notatka")

        # Check fractions
        zmieszane_1100_idx = HEADERS.index('zmieszane_1100') + 1
        self.assertEqual(ws.cell(row=2, column=zmieszane_1100_idx).value, 3)

        bio_idx = HEADERS.index('bio') + 1
        self.assertEqual(ws.cell(row=2, column=bio_idx).value, 1)

        # Check an empty fraction
        makulatura_1100_idx = HEADERS.index('makulatura_1100') + 1
        self.assertEqual(ws.cell(row=2, column=makulatura_1100_idx).value or '', '')
